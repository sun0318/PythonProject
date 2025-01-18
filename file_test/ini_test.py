'''
1.  *   读取ini格式的文件，并创建一个excel文件，且为每个节点创建一个sheet，然后将节点下的键值写入到excel中，按照如下格式。 <img src alt="image-20201218204922898" style />

        *   首行，字体白色 & 单元格背景色蓝色。
        *   内容均居中。
        *   边框。

'''
import configparser
from openpyxl import workbook
from openpyxl.styles import PatternFill, Font, Border

config = configparser.ConfigParser()
config.read('mysql.ini', encoding='utf-8')
ret = config.sections()
wb = workbook.Workbook()
# del wb["sheet"]

fill = PatternFill("solid", fgColor="3300ff")
font = Font(name="微软雅黑", color="ffffff", underline="single")
title_list = {"A1": "键", "B1": "值"}
for items in ret:
    sheet = wb.create_sheet(items)
    for position, title in title_list.items():
        cell = sheet[position]
        cell.value = title
    item_list = config.items(items)
    print(item_list)
    row_count=2
    for group in item_list:
        for index ,value in enumerate(group,1):
            sheet.cell(row_count,index).value=value
        row_count+=1
    # for i in range(len(item_list)-1):
    #     x = "A{}".format(i+2)
    #     y = "B{}".format(i+2)
    #     sheet[x].value = item_list[i][0]
    #     sheet[y].value = item_list[i][1]
wb.save("ini.xlsx")
