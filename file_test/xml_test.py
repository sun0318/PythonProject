import requests
from xml.etree import ElementTree as ET
from openpyxl import workbook

count = 0
wb = workbook.Workbook()
# 默认打开sheet
del wb['Sheet']
wb.active = 0
while True:
    city = input("请输入城市（Q/q退出）：")
    if city.upper() == "Q":
        break
    url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(city)
    res = requests.get(url=url)
    # print(res.text)
    # 1.提取XML格式中的数据
    root = ET.XML(res.text)
    print(len(root))
    # 2.为每个城市创建一个sheet，并将获取的xml格式中的数据写入到excel中。

    sheet = wb.create_sheet(city, count)
    for i in range(len(root)):
        cell = sheet.cell(i + 1, 1)
        cell.value = root[i].text
        # print(root[i].text)
    count = count + 1
wb.save("city.xlsx")